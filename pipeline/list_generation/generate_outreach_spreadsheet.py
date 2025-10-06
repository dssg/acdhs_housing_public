import pandas as pd
from sqlalchemy import create_engine
import numpy as np
from openpyxl.styles import Protection
import openpyxl
from pathlib import Path
import argparse
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
import random
import warnings
import datetime
warnings.simplefilter(action = "error", category=FutureWarning)
pd.options.mode.chained_assignment = 'raise'
import sys
import os
import logging
from pipeline.utils.project_constants import LOGS_PATH

# set up logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
now = datetime.datetime.now()
date_time = now.strftime("%Y-%m-%d_%H:%M:%S")
fh = logging.FileHandler(f'{LOGS_PATH}/spreadsheet_run_{date_time}.log', mode='w')
fh.setFormatter(formatter)
logger.addHandler(fh)
ch = logging.StreamHandler()
ch.setFormatter(formatter)
logger.addHandler(ch)

orgs = ["PSH_RRH", "CYF", "RentHelp", "OCS", "MDJ", "ACTION"]
prohibited_assignments = {}
first_time_names = {}
for org in orgs:
    prohibited_assignments[org] = set()
    first_time_names[org] = set()
status_columns = ["Contact status", "Court status", "Rental assistance status", "Did this help tenant?"]

completed_cases = set()
action_assignments = set()

def generate_spreadsheet(spreadsheet_date, pfwd_as_of_date, model_id, out_filename, write_to_db, rct, random_seed, sample_seeds):
    random.seed(random_seed) # set seed
    spreadsheet_date = pd.to_datetime(spreadsheet_date)
    engine = create_engine("postgresql:///?service=acdhs_housing")
    get_complete_cases(engine)
    incomplete_cases = get_incomplete_cases(engine) # some of these belong to action -- we need to get fresh data (program enrollments etc) to dole these out to the right outreach groups
    get_inactive_cases(engine, incomplete_cases) # none of these belong to action -- this should be a subset of incomplete_last_week bc if cases are complete, it doesn't matter that they're inactive
    # all inactive cases should be added to action's workload, they will definitely be present in incomplete_cases, if they are still in the predict_forward list they should be eliminated for already having been generated in a previous list
    # add up-to-date context data to incomplete cases
    incomplete_cases = add_context_data(engine, incomplete_cases, spreadsheet_date) 
    if rct:
        full_list = fill_capacity_with_predict_fwd_rct(engine, incomplete_cases, pfwd_as_of_date, model_id, spreadsheet_date, write_to_db, random_seed, sample_seeds) 
    else:
        full_list = fill_capacity_with_predict_fwd(engine, incomplete_cases, 30, pfwd_as_of_date, model_id, spreadsheet_date) # TODO k is hardcoded here but also we're not using this function
    full_list = add_notes_status_rank_history(full_list, engine, spreadsheet_date)
    # check: max and min scores 
    max_score = full_list['score'].max()
    min_score = full_list['score'].min()
    logging.info(f"Max score: {max_score}")
    logging.info(f"Min score: {min_score}")
    if max_score == min_score:
        logging.warning(f"WARNING: max and min scores are the same")
    # save output
    full_list = format_spreadsheet(full_list)
    full_list = full_list.sample(frac=1, random_state=random_seed).reset_index(drop=True)
    with pd.ExcelWriter(out_filename, engine='openpyxl') as writer:
        full_list.to_excel(writer, sheet_name=f"full_list", index=False, freeze_panes=(1,0))
        # format column width of full sheet
        workbook = writer.book
        sheet = workbook["full_list"]
        for col in range(1, len(full_list.columns) + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            col_name = full_list.columns[col-1]
            if col_name == "history_desc":
                sheet.column_dimensions[col_letter].width = 30
            if col_name in ['previous_notes', 'history_desc']: 
                continue
            column_width = max(full_list[col_name].astype(str).map(len).max(), len(col_name))
            sheet.column_dimensions[col_letter].width = column_width + 1
        # lock full sheet 
        sheet.protection.sheet = True
        full_list = full_list.loc[full_list['dod'].isnull()] # remove individuals who have a recorded date of death
        for outreach_org in orgs:
            format_outreach_spreadsheet(writer, full_list, outreach_org, out_filename)
    # write to db
    if write_to_db:
        full_list['db_recorded_timestamp'] = pd.Timestamp.now()
        full_list.rename(columns={'assign_ACTION': 'assign_action', 'assign_CYF': 'assign_cyf', 'assign_OCS': 'assign_ocs', 'assign_MDJ': 'assign_mdj', 'assign_RentHelp': 'assign_renthelp', 'assign_PSH_RRH': 'assign_psh_rrh'}, inplace=True)
        full_list.to_sql('outreach_lists', engine, index=False, if_exists='append', schema='acdhs_production')
        logging.info("wrote to db table outreach_lists")
    # check: most recent filingdt
    max_filing_dt = pd.to_datetime(full_list['filingdt'].max())
    logging.info(f"Latest filing date: {max_filing_dt.date()}")
    if spreadsheet_date - max_filing_dt > datetime.timedelta(days=3):
        logging.warning(f"WARNING: max_filing_dt older than 3 days ago ({spreadsheet_date - max_filing_dt})")

def map_cases_to_outreach(cases):
    for outreach in orgs:
        cases[f"assign_{outreach}"] = cases.apply(map_case_to_outreach, outreach_org=outreach, axis=1)
    cases["outreach_list"] = cases.apply(create_outreach_list, axis=1)
    # indicate unassigned cases
    cases['has_assignment'] = False
    cases.loc[(cases['assign_ACTION']==1)|(cases['assign_CYF']==1)|(cases['assign_OCS']==1)|(cases['assign_RentHelp']==1)|(cases['assign_MDJ']==1)|(cases['assign_PSH_RRH']==1), 'has_assignment'] = True
    cases.reset_index(drop=True, inplace=True)
    return cases

def create_outreach_list(case_row):
    case_id = (case_row['client_hash'], case_row['filingdt'].strftime('%Y-%m-%d'))
    str_list = []
    if case_row['assign_ACTION']==1:
        str_list.append('ACTION')
    if case_row['assign_CYF']==1:
        str_list.append('CYF')
    if case_row['assign_MDJ']==1:
        str_list.append('MDJ')
    if case_row['assign_OCS']==1:
        str_list.append('OCS')
    if case_row['assign_PSH_RRH']==1:
        str_list.append('PSH_RRH')
    if case_row['assign_RentHelp']==1:
        str_list.append('RentHelp')
    if case_id in completed_cases:
        str_list.append('completed case')
    return ",".join(str_list)

def map_case_to_outreach(case_row, outreach_org):
    case_id = (case_row['client_hash'], case_row['filingdt'].strftime('%Y-%m-%d'))
    prohibited_assignments_org = prohibited_assignments[outreach_org]
    if case_id in prohibited_assignments_org:
        return 0
    if case_id in completed_cases:
        # completed case is not added to any organizations list -- this overrules all other records
        return 0 
    if outreach_org == "PSH_RRH":
        if case_row['outreach_psh_rrh'] == 1:
            return 1
        else:
            return 0
    elif outreach_org == "CYF":
        if case_row['outreach_cyf'] == 1 and case_row['client_hash'] not in prohibited_assignments_org:
            # CYF has a special check for the client_hash with no particular filingdt -- these people are no longer eligible for CYF even if they're listed as enrolled
            return 1
        else:
            return 0
    elif outreach_org == "RentHelp":
        if case_row['docketno'].startswith(('MJ-05228', 'MJ-05312', 'MJ-05242', 'MJ-05240')) and pd.isnull(case_row['dispositiondt']):
            # dispositiondt isn't recorded until the disposition hearing happens -- there will never be a future disposition date in our data
            # TODO future considerations: for now RentHelp individuals don't appear on ACTION's list until the disposition hearing is past OR they're inactive for two weeks -- any initial tenants who go to RentHelp's list aren't added to ACTION until those conditions are met
            return 1
        else:
            return 0 
    elif outreach_org == "OCS":
        if (case_row['outreach_ocs_x'] == 1 or case_row['outreach_ocs_y'] == 1):
            return 1
        else:
            return 0
    elif outreach_org == "MDJ":
        if pd.isnull(case_row['dispositiondt']):
            return 1
        else:
            return 0
    elif outreach_org == 'ACTION':
        if case_row['assign_PSH_RRH'] == 1:
            # ACTION can't pay the rent for anyone who's currently in a PSH/RRH program
            return 0
        if case_id in action_assignments:
            return 1
        elif case_row['assign_CYF'] == 0 and case_row['assign_PSH_RRH'] == 0 and case_row['assign_RentHelp'] == 0 and case_row['assign_OCS'] == 0:
            # need to check assign_ columns rather than outreach_ columns because they contain all the logic for completed/prohibited/etc cases
            return 1
        else:
            return 0
        

def assign_tier(row): 
    if row['rank_within_group'] <= 30:
        tier = "top"
    elif row['rank_within_group'] <= 300:
        tier = "middle"
    else:
        tier = "bottom"
    return tier

def fill_capacity_with_predict_fwd_rct(engine, incomplete_cases, pfwd_as_of_date, model_id, spreadsheet_date, write_to_db, random_seed, sample_seeds):
    # check distribution of incomplete cases 
    incomplete_cases_assignments = map_cases_to_outreach(incomplete_cases)
    incomplete_cases_assignments = incomplete_cases_assignments.loc[incomplete_cases_assignments['has_assignment']==True]
    logging.info("Assignment count from incomplete cases:")
    for org in orgs:
        col_name = f"assign_{org}"
        logging.info(f"{org}: {incomplete_cases_assignments[col_name].sum()} cases")
    logging.info(f"Now adding cases from predict_forward list or cases previously sampled into treatment group")
    # get cases from this week's predict_forward list
    pf_list = get_predict_forward_list(engine, model_id, pfwd_as_of_date) # full pfwd list (full cohort)
    orig_cohort_size = len(pf_list)
    logging.info(f"predict forward cohort size: {len(pf_list)}")
    pf_list_sub = pf_list.loc[~pf_list['client_hash'].isin(incomplete_cases_assignments['client_hash'])] # drop duplicates
    pf_list_final = pf_list_sub.dropna(subset=['mci_uniq_id'])
    pf_list_final = add_context_data(engine, pf_list_final, spreadsheet_date)
    pf_list_outreach_assignments = map_cases_to_outreach(pf_list_final) 
    pf_list_valid = pf_list_outreach_assignments.loc[pf_list_outreach_assignments['has_assignment']==True] # these are all the cases from the predict_forward list that have valid outreach assignents and aren't completed
    pf_list_valid.set_index('mci_uniq_id', inplace=True)
    logging.info(f"predict forward list, valid/not completed cases: {len(pf_list_valid)}") 
    # get c/t assignments from db
    q = f"""
        select 
            distinct on (mci_uniq_id)
            mci_uniq_id,
            client_hash,
            entity_id,
            model_id, 
            as_of_date,
            ct_group,
            group_assignment_date,
            selected_for_outreach,
            tier,
            outreach_selection_date,
            score, 
            rank_abs_no_ties,
            assignment_seed,
            sample_seed1,
            sample_seed2,
            sample_seed3
        from acdhs_production.control_treatment_assignment
        order by mci_uniq_id, db_recorded_timestamp desc
        """
    all_past_assignments = pd.read_sql(q, engine, parse_dates=['as_of_date', 'group_assignment_date', 'outreach_selection_date'])
    all_past_assignments.set_index('mci_uniq_id', inplace=True)
    logging.info(f"all past assignments from acdhs_production.control_treatment_assignment: {len(all_past_assignments)}")
    # find mci_uniq_ids in this week's cohort (those with valid outreach assignment only) that already have previous C/T assignments
    cohort_past_assignments = all_past_assignments[all_past_assignments.index.isin(pf_list_valid.index)]
    # update with new columns from this week's cohort
    cohort_past_assignments = cohort_past_assignments.join(pf_list_valid[['client_hash', 'entity_id', 'model_id', 'as_of_date', 'score', 'rank_abs_no_ties']], how='left', lsuffix='_previous', rsuffix='_new')
    cohort_past_assignments.drop(['client_hash_previous', 'entity_id_previous', 'model_id_previous', 'as_of_date_previous', 'score_previous', 'rank_abs_no_ties_previous'], axis=1, inplace=True)
    cohort_past_assignments = cohort_past_assignments.rename(columns={'client_hash_new': 'client_hash', 'entity_id_new': 'entity_id', 'model_id_new': 'model_id', 'as_of_date_new': 'as_of_date', 'score_new': 'score', 'rank_abs_no_ties_new': 'rank_abs_no_ties'})
    logging.info(f"past assignments for this week's (valid) cohort from acdhs_production.control_treatment_assignment: {len(cohort_past_assignments)}")
    if len(cohort_past_assignments) > 0:
        cohort_past_assignments.reset_index(inplace=True)
        cohort_past_assignments = add_context_data(engine, cohort_past_assignments, spreadsheet_date)
        cohort_past_assignments = map_cases_to_outreach(cohort_past_assignments) # nothing should get dropped here bc we've already restricted with pf_list_valid 
        logging.info(f"are there any cohort_past_assignments with no assignment? {len(cohort_past_assignments.loc[cohort_past_assignments['has_assignment']==False])}")
        cohort_past_assignments = cohort_past_assignments.loc[cohort_past_assignments['has_assignment']==True]
        cohort_past_assignments.set_index('mci_uniq_id', inplace=True)
    logging.info(f"past assignments valid for selection and not completed cases (is this the same as the previous len count?): {len(cohort_past_assignments)}")
    # cases from predict forward that don't have control/treatment assignment yet (they already have outreach assignments from pf_list_valid)
    unassigned_list = pf_list_valid.copy()
    unassigned_list = unassigned_list[~unassigned_list.index.isin(all_past_assignments.index)]
    logging.info(f"predict forward not completed cases that don't have control/treatment assignments yet: {len(unassigned_list)}")
    # assign to control/treatment
    unassigned_list['ct_group'] = random.choices(["control", "treatment"], k=len(unassigned_list))
    unassigned_list['assignment_seed'] = random_seed
    logging.info(f"{len(unassigned_list.loc[unassigned_list['ct_group']=='control'])} cases added to control group")
    logging.info(f"{len(unassigned_list.loc[unassigned_list['ct_group']=='treatment'])} cases added to treatment group")
    unassigned_list['group_assignment_date'] = spreadsheet_date
    all_assignments_for_selection = pd.concat([cohort_past_assignments, unassigned_list])
    # now sample new outreach list for this week
    potential_treatment = all_assignments_for_selection.copy()
    potential_treatment = potential_treatment.loc[(potential_treatment['ct_group']=='treatment') & (potential_treatment['selected_for_outreach']!=True)]
    logging.info(f"{len(potential_treatment)} possible outreach candidates")
    potential_treatment.sort_values(by=['score', 'rank_abs_no_ties'], ascending=[False, True], inplace=True)
    # top sample
    top_rows = potential_treatment.iloc[:30,:] 
    top_sample = top_rows.sample(n=15, random_state=sample_seeds[0])
    potential_treatment.loc[potential_treatment.index.isin(top_sample.index), 'tier'] = 'top'
    # middle sample
    middle_rows = potential_treatment.iloc[30:300,:] 
    middle_sample = middle_rows.sample(n=10, random_state=sample_seeds[1])
    potential_treatment.loc[potential_treatment.index.isin(middle_sample.index), 'tier'] = 'middle'
    # bottom sample
    bottom_rows = potential_treatment.iloc[300:,:] 
    bottom_sample = bottom_rows.sample(n=5, random_state=sample_seeds[2])
    potential_treatment.loc[potential_treatment.index.isin(bottom_sample.index), 'tier'] = 'bottom'
    # full list
    full_list = pd.concat([top_sample, middle_sample, bottom_sample])
    potential_treatment.loc[potential_treatment.index.isin(full_list.index), 'selected_for_outreach'] = True
    potential_treatment.loc[potential_treatment.index.isin(full_list.index), 'outreach_selection_date'] = spreadsheet_date
    potential_treatment['sample_seed1'] = sample_seeds[0]
    potential_treatment['sample_seed2'] = sample_seeds[1]
    potential_treatment['sample_seed3'] = sample_seeds[2]
    # get ranks within group and tier assignments
    control_group = all_assignments_for_selection.copy()
    control_group = control_group.loc[control_group['ct_group']=='control']
    control_group.sort_values(by=['score', 'rank_abs_no_ties'], inplace=True, ascending=[False, True])
    control_group.reset_index(inplace=True)
    control_group.reset_index(inplace=True)
    control_group = control_group.rename(columns={"index":"rank_within_group"})
    control_group['rank_within_group'] = control_group['rank_within_group'] + 1
    logging.info("CONTROL")
    logging.info(control_group[['score', 'rank_abs_no_ties', 'rank_within_group']])
    potential_treatment.sort_values(by=['score', 'rank_abs_no_ties'], ascending=[False, True], inplace=True)
    potential_treatment.reset_index(inplace=True)
    potential_treatment.reset_index(inplace=True)
    potential_treatment = potential_treatment.rename(columns={"index":"rank_within_group"})
    potential_treatment['rank_within_group'] = potential_treatment['rank_within_group'] + 1
    logging.info("POTENTIAL TREATMENT")
    logging.info(potential_treatment[['score', 'rank_abs_no_ties', 'rank_within_group']])
    control_plus_potential_treatment = pd.concat([control_group, potential_treatment])
    control_plus_potential_treatment['tier'] = control_plus_potential_treatment.apply(assign_tier, axis=1)
    control_plus_potential_treatment = control_plus_potential_treatment.rename(columns={"index":"rank_within_group"})
    # control_plus_potential_treatment.reset_index(inplace=True)
    # now add any cases missing from cohort (already completed etc)
    missing_cases = pf_list.loc[~pf_list['client_hash'].isin(control_plus_potential_treatment['client_hash'])] 
    logging.info(f"number of missing cases: {len(missing_cases)}")
    updates = pd.concat([control_plus_potential_treatment, missing_cases])
    logging.info(f"updates: {len(updates)}")
    assign_cohort_size = len(updates)
    logging.info(f"assignments cohort size: {len(pf_list)}")
    if orig_cohort_size != assign_cohort_size:
        logging.error(f"ERROR: cohort size in assignments table ({assign_cohort_size}) does not match original cohort size ({orig_cohort_size})")
        sys.exit(0)
    # update db with full cohort assignments
    updates['db_recorded_timestamp'] = pd.Timestamp.now()
    updates = updates.drop(['prediction_date',
       'matter_id', 'docketno', 'filingdt',
       'dispositiondt', 'city_of_pgh', 'ofp_issue_dt', 'arrears_only',
       'judgment_for_landlord', 'entry_of_satisfaction_issued',
       'entry_of_satisfaction_issue_dt', 'entry_of_satisfaction_requested',
       'entry_of_satisfaction_request_dt', 'claimamount',
       'totaljudgmentamount', 'prev_ra_payment_tnumber', 'prev_ra_payment',
       'psh_rrh_start_dt', 'psh_rrh_end_dt', 'outreach_psh_rrh',
       'cyf_start_dt', 'cyf_end_dt', 'cyf_entity_type', 'cyf_entity_id',
       'outreach_cyf', 'hello_baby_contact_date', 'priority_provider',
       'outreach_ocs_x', 'family_center_contact_date', 'facility_name',
       'assigned_worker_name', 'contact_worker_name', 'outreach_ocs_y',
       'fname', 'mname', 'lname', 'name_suffix', 'cell_phone', 'home_phone',
       'work_phone', 'dod', 'prev_link_call', 'prev_link_screen',
       'prev_ra_app_tnumber', 'prev_ra_app_update', 'prev_ra_app_submitted',
       'erap_months_regular_table', 'erap_months', 'assign_PSH_RRH',
       'assign_CYF', 'assign_RentHelp', 'assign_OCS', 'assign_MDJ',
       'assign_ACTION', 'outreach_list'], axis=1)
    if write_to_db:
        updates.to_sql('control_treatment_assignment', engine, index=False, if_exists='append', schema='acdhs_production')
        logging.info("wrote to db table control_treatment_assignment")
    # full outreach list is incomplete cases from previous list plus the new outreach cases sampled from the predict_forward list/previous cases sampled for treatment
    full_list.reset_index(inplace=True)
    full_list.drop(['ct_group', 'group_assignment_date', 'selected_for_outreach',
       'tier', 'outreach_selection_date'], axis=1, inplace=True)
    logging.info(f"{len(full_list)} new cases from predict_forward cohort/previous cases sampled for treatment")
    logging.info(f"Added to {len(incomplete_cases_assignments)} incomplete cases from the previous week")
    full_list = pd.concat([incomplete_cases_assignments, full_list]).reset_index(drop=True)
    return full_list

def fill_capacity_with_predict_fwd(engine, incomplete_cases, k, pfwd_as_of_date, model_id, spreadsheet_date):
    # how many incomplete cases are assigned to ACTION
    incomplete_cases_assignments = map_cases_to_outreach(incomplete_cases)
    logging.info("Assignment count from incomplete cases:")
    for org in orgs:
        col_name = f"assign_{org}"
        logging.info(f"{org}: {incomplete_cases_assignments[col_name].sum()} cases")
    logging.info(f"Now adding {k} cases from predict_forward list")
    # get more cases from this week's predict_forward list
    pf_list = get_predict_forward_list(engine, model_id, pfwd_as_of_date) 
    pf_list_sub = pf_list.loc[~pf_list['client_hash'].isin(incomplete_cases_assignments['client_hash'])] # drop duplicates
    pf_list_final = pf_list_sub.dropna(subset=['mci_uniq_id'])
    pf_list_final = add_context_data(engine, pf_list_final, spreadsheet_date)
    pf_list_assignments = map_cases_to_outreach(pf_list_final)
    pf_list_fill = pf_list_assignments.iloc[:k,:]
    logging.info(f"{len(pf_list_fill)} new cases from predict_forward")
    full_list = pd.concat([incomplete_cases_assignments, pf_list_fill]).reset_index(drop=True)
    return full_list

# get full predict forward list (not limited to top k by rank)
def get_predict_forward_list(engine, model_id, as_of_date): 
    q = f"""
        select 
            mci_uniq_id,
            model_id, 
            client_hash,
            entity_id,
            as_of_date,
            prediction_date,
            score, 
            rank_abs_no_ties
        from acdhs_production.predictions 
        left join clean.mci_hash_mapping using(client_hash)
        where model_id={model_id} and as_of_date='{as_of_date}'
        """
    output = pd.read_sql(q, engine, parse_dates=['as_of_date', 'prediction_date'])
    # output = output.dropna(subset=['mci_uniq_id'])
    output = output.sort_values(by=['rank_abs_no_ties'])
    return output

def get_complete_cases(engine):
    # acdhs_production.outreach_lists: past cases only
    q = """
        select 
        distinct client_hash, filingdt, data_source, "Contact status", "Rental assistance status"
        from acdhs_production.outreach_contact_notes ocn join acdhs_production.outreach_lists using(client_hash, model_id, as_of_date, spreadsheet_date, prediction_date)
        where "Contact status"='Unable to contact' or "Rental assistance status" in ('Unable to contact', 'Not eligible/can''t help', 'Payment made/approved', 'Not in Allegheny County')
    """
    complete_df = pd.read_sql(q, engine, parse_dates=['filingdt'])
    # all cases marked complete by ACTION (EXCEPT unable to contact) are complete overall, won't be shown to any outreach group
    action_complete = complete_df.loc[(complete_df['data_source']=="ACTION") & (complete_df['Rental assistance status']!='Unable to contact')]
    completed_cases.update(list(zip(action_complete.client_hash, action_complete.filingdt.dt.strftime('%Y-%m-%d'))))
    # cases marked complete by some other outreach group are removed from that group's list, will be added to ACTION's list
    # This also includes ACTION unable to contact, which are removed from ACTION's list but NOT removed from other lists when relevant
    other_complete = complete_df.loc[(complete_df['data_source']!="ACTION") | ((complete_df['data_source']=="ACTION") & (complete_df['Rental assistance status']=='Unable to contact'))]
    for row in other_complete.itertuples():
        prohibited_assignments[row.data_source].add((row.client_hash, row.filingdt.strftime('%Y-%m-%d')))

def get_incomplete_cases(engine):
    # TODO somewhere here, check for cases that have since received payment/submitted app according to data warehouse -- remove them as complete
    # get incomplete cases from last week (acdhs_production.outreach_lists: past cases only)
    q = """
        with max_date as (
            select max(spreadsheet_date) as max_spreadsheet_date from acdhs_production.outreach_contact_notes
        )
        select 
            distinct on (client_hash, filingdt)
            ocn.client_hash, mci_uniq_id, ocn.model_id, ocn.as_of_date, ocn.spreadsheet_date, ocn.prediction_date, filingdt, matter_id, data_source, ocn."Attempt 1 notes", ocn."Attempt 2 notes", ocn."Attempt 3 notes", ocn."Contact status", ocn."Court status", ocn."Rental assistance status", ocn."Did this help tenant?"
        from acdhs_production.outreach_contact_notes ocn 
        join acdhs_production.outreach_lists using(client_hash, model_id, as_of_date, spreadsheet_date, prediction_date)
        join max_date on ocn.spreadsheet_date = max_date.max_spreadsheet_date
        where ("Contact status" is null or "Contact status"!='Unable to contact') and ("Rental assistance status" is null or "Rental assistance status" not in ('Unable to contact', 'Not eligible/can''t help', 'Payment made/approved', 'Not in Allegheny County'))
    """
    incomplete_df = pd.read_sql(q, engine, parse_dates=['as_of_date', 'spreadsheet_date', 'prediction_date', 'max_spreadsheet_date', 'filingdt'])
    # all cases that were on ACTION's list last week should remain on ACTION's list
    action_incomplete_df = incomplete_df.loc[incomplete_df['data_source']=='ACTION']
    action_incomplete_case_ids = set(tuple(zip(action_incomplete_df['client_hash'], action_incomplete_df['filingdt'].dt.strftime('%Y-%m-%d'))))
    action_assignments.update(action_incomplete_case_ids)
    return incomplete_df

def get_inactive_cases(engine, incomplete_last_week):
    # get cases that have been inactive for two weeks
    # first, get cases from three weeks ago (to compare to one week ago -- two week gap))
    # acdhs_production.outreach_lists: past cases only
    q = """
        select 
            distinct on (client_hash, filingdt)
            client_hash, filingdt, data_source, ocn."Attempt 1 notes", ocn."Attempt 2 notes", ocn."Attempt 3 notes", ocn."Contact status", ocn."Court status", ocn."Rental assistance status", ocn."Did this help tenant?"
        from acdhs_production.outreach_contact_notes ocn join acdhs_production.outreach_lists using(client_hash, model_id, as_of_date, spreadsheet_date, prediction_date)
        where ocn.spreadsheet_date=(
            SELECT DISTINCT spreadsheet_date FROM acdhs_production.outreach_contact_notes ORDER BY spreadsheet_date DESC LIMIT 1 OFFSET 2
        ) and data_source != 'ACTION';
    """
    prev_df = pd.read_sql(q, engine, parse_dates=['filingdt'])
    # find ids that existed in the list 3 weeks ago and 1 week ago (for the same data source)
    inactive_case_candidates = incomplete_last_week.merge(prev_df, how='left', on=['client_hash', 'filingdt', 'data_source']) 
    # drop any with data_source='PSH_RRH' because those can't be sent to ACTION
    inactive_case_candidates = inactive_case_candidates.loc[inactive_case_candidates['data_source']!='PSH_RRH']
    potentially_inactive_cases = set(list(zip(inactive_case_candidates['client_hash'], inactive_case_candidates['filingdt'].dt.strftime('%Y-%m-%d'))))
    # of those ids, which had updated notes/status => active case?
    inactive_case_candidates = inactive_case_candidates.fillna('')
    active_cases = inactive_case_candidates.loc[
        (inactive_case_candidates['Attempt 1 notes_x']!=inactive_case_candidates['Attempt 1 notes_y']) |
        (inactive_case_candidates['Attempt 2 notes_x']!=inactive_case_candidates['Attempt 2 notes_y']) |
        (inactive_case_candidates['Attempt 3 notes_x']!=inactive_case_candidates['Attempt 3 notes_y']) | 
        (inactive_case_candidates['Contact status_x']!=inactive_case_candidates['Contact status_y']) | 
        (inactive_case_candidates['Court status_x']!=inactive_case_candidates['Court status_y']) | 
        (inactive_case_candidates['Rental assistance status_x']!=inactive_case_candidates['Rental assistance status_y']) | 
        (inactive_case_candidates['Did this help tenant?_x']!=inactive_case_candidates['Did this help tenant?_y'])
        ]
    active_case_ids = set(tuple(zip(active_cases['client_hash'], active_cases['filingdt'].dt.strftime('%Y-%m-%d'))))
    # add the inactive cases (NOT an active case from week n-3 to week n-1 for any data source)
    # if at least one data source updated status/notes, the case is NOT inactive
    inactive_cases = potentially_inactive_cases - active_case_ids
    action_assignments.update(inactive_cases - completed_cases)

def add_context_data(engine, cases_df, spreadsheet_date):
    cases_df = get_eviction_data(cases_df, engine, spreadsheet_date)
    cases_df = get_ra_crrp_payments(cases_df, engine, spreadsheet_date)
    cases_df = get_housing_program_data(cases_df, engine, spreadsheet_date)
    cases_df = get_current_cyf(cases_df, engine, spreadsheet_date)
    cases_df = get_hello_baby_involvement(cases_df, engine, spreadsheet_date)
    cases_df = get_family_center_involvement(cases_df, engine, spreadsheet_date)
    cases_df = get_client_info(cases_df, engine)
    cases_df = get_death_dates(cases_df, engine)
    cases_df = get_link_calls(cases_df, engine, spreadsheet_date)
    cases_df = get_link_screening(cases_df, engine, spreadsheet_date)
    cases_df = get_ra_crrp_submission(cases_df, engine, spreadsheet_date)
    cases_df = get_erap_months(cases_df, engine)
    return cases_df

# make a spreadsheet version for each individual outreach group
def format_outreach_spreadsheet(writer, matrix, outreach_type, out_filename):
    cols_to_drop = ['assign_PSH_RRH', 'psh_rrh_start_dt', 'psh_rrh_end_dt', 'assign_CYF', 'cyf_start_dt', 'cyf_end_dt', 'cyf_entity_type', 'cyf_entity_id', 'assign_OCS', 'hello_baby_contact_date', 'priority_provider', 'family_center_contact_date', 'facility_name', 'assigned_worker_name', 'contact_worker_name', 'matter_id', 'dod', 'entity_id', 'assign_ACTION', 'assign_MDJ', 'assign_RentHelp']
    for org in orgs:
        for status_col in status_columns:
            col_to_check = f"{org} {status_col}"
            if org == outreach_type:
                matrix = matrix.rename(columns={col_to_check: status_col})
            else:
                cols_to_drop.append(col_to_check)
    if outreach_type == "PSH_RRH":
        matrix = matrix.loc[matrix['assign_PSH_RRH']==1]
        cols_to_drop.remove('psh_rrh_start_dt')
        cols_to_drop.remove('psh_rrh_end_dt')
    elif outreach_type == "CYF": 
        matrix = matrix.loc[matrix['assign_CYF']==1]
        cols_to_drop.remove('cyf_start_dt')
        cols_to_drop.remove('cyf_entity_type')
        cols_to_drop.remove('cyf_entity_id')
    elif outreach_type == "RentHelp": 
        matrix = matrix.loc[matrix['assign_RentHelp']==1]
    elif outreach_type == "OCS":
        matrix = matrix.loc[matrix['assign_OCS']==1]
        cols_to_drop.remove('hello_baby_contact_date')
        cols_to_drop.remove('priority_provider')
        cols_to_drop.remove('family_center_contact_date')
        cols_to_drop.remove('facility_name')
        cols_to_drop.remove('assigned_worker_name')
        cols_to_drop.remove('contact_worker_name')
    elif outreach_type == "MDJ":
        matrix = matrix.loc[matrix['assign_MDJ']==1]
    elif outreach_type == "ACTION": 
        matrix = matrix.loc[matrix['assign_ACTION']==1]
    unlocked_columns = ['Attempt 1 notes', 'Attempt 2 notes', 'Attempt 3 notes', "Contact status",
            "Court status",
            "Rental assistance status",
            "Did this help tenant?"]
    matrix = matrix.drop(cols_to_drop, axis=1)
    matrix = matrix.rename(columns={'prev_link_call': 'previous_link_call', 'prev_link_screen': 'previous_link_housing_screen', 'prev_ra_app_update': 'previous_rental_assistance_app_updated', 'prev_ra_app_submitted': 'previous_rental_assistance_app_submitted', 'prev_ra_app_tnumber': 'previous_rental_assistance_app_Tnumber', 'prev_ra_payment': 'previous_rental_assistance_payment', 'prev_ra_payment_tnumber': 'previous_rental_assistance_payment_tnumber', 'history_desc': 'Case history'})
    matrix = matrix.style.apply(highlight_new_names, org=outreach_type, axis=1)
    matrix.to_excel(writer, index=False, sheet_name=outreach_type, freeze_panes=(1,11))
    # Get the workbook and active sheet
    workbook = writer.book
    sheet = workbook[outreach_type]
    # Loop through each column and lock the cells if it's in locked_columns
    for col in range(1, len(matrix.columns) + 1):
        col_name = matrix.columns[col-1]
        col_letter = openpyxl.utils.get_column_letter(col)
        if matrix.columns[col - 1] in unlocked_columns:
            for cell in sheet[col_letter]:
                cell.protection = Protection(locked=False)
        column_width = max(matrix.data[col_name].astype(str).map(len).max(), len(col_name))
        sheet.column_dimensions[col_letter].width = column_width + 1
        if col_name == "name_suffix":
            sheet.column_dimensions[col_letter].width = 5
        if col_name == "previous_notes":
            sheet.column_dimensions[col_letter].width = 30
        if col_name == "Case history":
            sheet.column_dimensions[col_letter].width = 30
        if col_name == "Contact status":
            sheet.column_dimensions[col_letter].width = 13
        if col_name == "Court status":
            sheet.column_dimensions[col_letter].width = 16
        if col_name == "Rental assistance status":
            sheet.column_dimensions[col_letter].width = 21
        if col_name == "Did this help tenant?":
            sheet.column_dimensions[col_letter].width = 17
    # Protect the worksheet to enforce cell protection
    # Emily Bengel (OCS) asked that we not lock the sheet so she can filter by family center, Samantha Murphy (MDJ) asked for the same
    if outreach_type != "OCS" and outreach_type != "MDJ": 
        sheet.protection.sheet = True
    # annotation columns
    outcome_cols = [
        {'col_letter': 'F', 'options': {'Provider contacted': 'green', 'Attempted': 'yellow', 'Unable to contact': 'red', 'Contacted': 'green', 'No need to contact': 'green'}, 'additional_rules': {'Reached': 'green'}},
        {'col_letter': 'G', 'options': {"Didn't appear": 'red', "Appeared & didn't make contact": 'red', "Appeared & made contact": 'green'}},
        {'col_letter': 'H', 'options': {'Unable to contact': 'red', "Not eligible/can't help": 'red', 'Application in progress': 'yellow', 'Payment made/approved': 'green', 'Not in Allegheny County': 'gray'}},
        {'col_letter': 'I', 'options': {'Yes--created new application': 'green', 'Yes--prioritized existing application': 'green', 'No': 'normal'}, 'additional_rules': {'Yes': 'green'}},
    ]
    check_colors = {
        'green': {'font': Font(color="136e16"), 'fill': PatternFill(bgColor="c3ebcb")},
        'red': {'font': Font(color="9C0006"), 'fill': PatternFill(bgColor="FFC7CE")},
        'yellow': {'font': Font(color="a05a09"), 'fill': PatternFill(bgColor="ffec9c")},
        'blue': {'font': Font(color="285c9c"), 'fill': PatternFill(bgColor="a8ccec")},
        'normal': {'font': Font(color="000000"), 'fill': PatternFill(bgColor="ffffff")},
        'gray': {'font': Font(color="ffffff"), 'fill':  PatternFill(bgColor="a8a4a4")}
    }
    for outcome in outcome_cols:
        # create data validation options
        list_options = ['--'] + list(outcome['options'].keys())
        dv = DataValidation(type="list", formula1=f'"{",".join(list_options)}"', showDropDown=False, allow_blank=True) # To see the drop down arrow, use "showDropDown=False"
        dv.error ='Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'
        dv.showInputMessage = True
        dv.showErrorMessage = True
        sheet.add_data_validation(dv)
        dv.add(f"{outcome['col_letter']}2:{outcome['col_letter']}201")
        # add conditional formatting color rules
        format_outcomes = dict(outcome['options'])
        if 'additional_rules' in outcome:
            format_outcomes.update(outcome['additional_rules'])
        # for option_name in outcome['options'] + outcome['additional_rules']:
        for option_name in format_outcomes:
            if option_name != 'No':
                dxf = DifferentialStyle(font=check_colors[format_outcomes[option_name]]['font'], fill=check_colors[format_outcomes[option_name]]['fill']) 
                rule = Rule(type="containsText", operator="containsText", text=option_name, dxf=dxf)
                rule.formula = [f'NOT(ISERROR(SEARCH("{option_name}",{outcome["col_letter"]}2)))']
                sheet.conditional_formatting.add(f"{outcome['col_letter']}2:{outcome['col_letter']}201", rule)
    # Save the Excel file
    workbook.save(out_filename)

def highlight_new_names(r, org):
    if r.client_hash in first_time_names[org]:
        colors = ["background-color: #9abddb"] * 4 + [''] * (len(r) - 4)
    else:
        colors = [''] * len(r)
    return colors

# get most recent eviction 
def get_eviction_data(matrix, engine, spreadsheet_date):
    client_hash_list = "','".join(matrix['client_hash'])
    client_hash_list = f"'{client_hash_list}'"
    q = f"""
        select 
            matter_id, 
            ev.docketno,
            filingdt,
            ev.dispositiondt,
            city_of_pgh_flag as city_of_pgh, 
            ofp_issue_dt,
            grantpossessionjudgmentnotsat as arrears_only,
            judgement_for_landlord as judgment_for_landlord,
            entry_of_satisfaction_issued,
            entry_of_satisfaction_issue_dt,
            entry_of_satisfaction_requested,
            entry_of_satisfaction_request_dt,
            hashed_mci_uniq_id as client_hash,
            claimamount,
            totaljudgmentamount
        from clean.eviction ev join clean.eviction_client_matches ecm using (matter_id)
        where ecm.hashed_mci_uniq_id in ({client_hash_list})
    """
    evictions = pd.read_sql(q, engine, parse_dates=['filingdt', 'dispositiondt', 'ofp_issue_dt'])
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['client_hash']]
    evictions_id = matrix_tmp.merge(evictions, how='inner', on='client_hash') # inner join: if we can't find eviction data, ignore this case
    evictions_id = evictions_id.loc[evictions_id['filingdt'] < spreadsheet_date]
    evictions_id = evictions_id.sort_values(by=['filingdt'], ascending=False)
    # set defaults when dispositiondt is unknown
    evictions_id.loc[evictions_id['dispositiondt'].isna(), 'judgment_for_landlord'] = ''
    evictions_id.loc[evictions_id['dispositiondt'].isna(), 'arrears_only'] = ''
    # only entry_of_satisfaction=True values are meaningful
    evictions_id.loc[evictions_id['entry_of_satisfaction_issued']==False, 'entry_of_satisfaction_issued'] = ''
    evictions_id.loc[evictions_id['entry_of_satisfaction_requested']==False, 'entry_of_satisfaction_requested'] = ''
    evictions_id = evictions_id.drop_duplicates(subset=['client_hash'])
    if 'filingdt' in matrix.columns:
        output = matrix.merge(evictions_id, how='inner', on=['client_hash','filingdt', 'matter_id'])
    else:
        output = matrix.merge(evictions_id, how='inner', on=['client_hash'])
    return output

def get_erap_months(matrix, engine):
    mci_ids = matrix.loc[~matrix['mci_uniq_id'].isnull()]['mci_uniq_id']
    mci_id_list = ",".join(mci_ids.astype(int).astype(str))
    q = f"""
        select 
            mci_uniq_id,
            rent_months as erap_months_regular_table
        from clean.erap_months_regular_table emrt where mci_uniq_id in ({mci_id_list})
    """
    df = pd.read_sql(q, engine)
    matrix = matrix.merge(df, how='left', on='mci_uniq_id')
    q = f"""
        select 
            mci_uniq_id,
            rent_months as erap_months
        from clean.erap_months_reporting_table emrt where mci_uniq_id in ({mci_id_list})
    """
    df = pd.read_sql(q, engine)
    matrix = matrix.merge(df, how='left', on='mci_uniq_id')
    return matrix

# get closest RA payments before and after spreadsheet_date (limit 1 year)
def get_ra_crrp_payments(matrix, engine, spreadsheet_date):
    q = f"""
        select
            mci_uniq_id,
            crrp_app_id,
            last_payment_dt as ra_last_payment_dt
        from clean.rental_assistance_payment_status 
    """
    payment_status = pd.read_sql(q, engine, parse_dates=['ra_last_payment_dt'])
    # find previous payment
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['mci_uniq_id']]
    output_prev = matrix_tmp.merge(payment_status, how='left', on='mci_uniq_id')
    output_prev = output_prev.drop(output_prev[output_prev['ra_last_payment_dt']>spreadsheet_date].index) # drop payments after spreadsheet_date (current date)
    output_prev = output_prev.drop(output_prev[output_prev['ra_last_payment_dt']<spreadsheet_date-pd.DateOffset(years=1)].index) # drop payments before spreadsheet_date (current date) - 1 year
    output_prev = output_prev.sort_values(by=['ra_last_payment_dt'], ascending=False)
    output_prev = output_prev.drop_duplicates(subset=['mci_uniq_id'])
    output_prev = output_prev.dropna(subset=['ra_last_payment_dt']) # drop people who have no previous payment under these rules
    output_prev = output_prev.rename(columns={'ra_last_payment_dt': 'prev_ra_payment', 'crrp_app_id': 'prev_ra_payment_tnumber'})
    output_prev['prev_ra_payment_tnumber'] = 'T' + output_prev['prev_ra_payment_tnumber'].astype(int).astype(str)
    # merge with full list
    output_prev['mci_uniq_id'] = output_prev['mci_uniq_id'].astype(int)
    matrix = matrix.merge(output_prev, how='left', on=['mci_uniq_id'])
    return matrix

def get_ra_crrp_submission(matrix, engine, spreadsheet_date):
    client_hash_list = "','".join(matrix['client_hash'])
    client_hash_list = f"'{client_hash_list}'"
    q = f"""
        select
            hashed_mci_uniq_id as client_hash, 
            crrp_app_id,
            --case when lst_updt_dt > lst_review_dt then lst_updt_dt else lst_review_dt end as ra_lst_updt_dt,
            coalesce(case when lst_updt_dt > lst_review_dt then lst_updt_dt else lst_review_dt end, lst_updt_dt, lst_review_dt) as ra_lst_updt_dt,
            submitted_dt as ra_app_submitted_dt
        from clean.rental_assistance_household rah join 
        clean.rental_assistance_application raa using(crrp_app_id)
        where hashed_mci_uniq_id in ({client_hash_list})
    """
    ra_df = pd.read_sql(q, engine, parse_dates=['ra_app_submitted_dt'])
    ra_df['ra_lst_updt_dt'] = pd.to_datetime(ra_df['ra_lst_updt_dt'])
    # find previous payment
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['client_hash']]
    output_prev = matrix_tmp.merge(ra_df, how='left', on='client_hash')
    output_prev = output_prev.drop(output_prev[output_prev['ra_lst_updt_dt']>spreadsheet_date].index) # drop submissions after spreadsheet_date (current date)
    output_prev = output_prev.drop(output_prev[output_prev['ra_lst_updt_dt']<spreadsheet_date-pd.DateOffset(months=6)].index) # drop submissions before spreadsheet_date (current date) - 6 months
    output_prev = output_prev.sort_values(by=['ra_lst_updt_dt'], ascending=False)
    output_prev = output_prev.drop_duplicates(subset=['client_hash'])
    output_prev = output_prev.dropna(subset=['ra_lst_updt_dt']) # drop people who have no previous payment under these rules
    output_prev = output_prev.rename(columns={'ra_lst_updt_dt': 'prev_ra_app_update', 'ra_app_submitted_dt': 'prev_ra_app_submitted', 'crrp_app_id': 'prev_ra_app_tnumber'})
    output_prev['prev_ra_app_tnumber'] = 'T' + output_prev['prev_ra_app_tnumber'].astype(int).astype(str)
    # merge with full list
    matrix = matrix.merge(output_prev, how='left', on=['client_hash'])
    return matrix

def get_link_calls(matrix, engine, spreadsheet_date):
    client_hash_list = "','".join(matrix['client_hash'])
    client_hash_list = f"'{client_hash_list}'"
    q = f"""
        select
            hashed_mci_uniq_id as client_hash, 
            referral_date as link_call_date
        from clean.link_coordinated_entry
        where hashed_mci_uniq_id in ({client_hash_list})
    """
    link_df = pd.read_sql(q, engine, parse_dates=['link_call_date'])
    # find previous call
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['client_hash']]
    output_prev = matrix_tmp.merge(link_df, how='left', on='client_hash')
    output_prev = output_prev.drop(output_prev[output_prev['link_call_date']>spreadsheet_date].index) # drop calls after spreadsheet_date (current date)
    output_prev = output_prev.drop(output_prev[output_prev['link_call_date']<spreadsheet_date-pd.DateOffset(months=6)].index) # drop calls before spreadsheet_date (current date) - 6 months
    output_prev = output_prev.sort_values(by=['link_call_date'], ascending=False)
    output_prev = output_prev.drop_duplicates(subset=['client_hash'])
    output_prev = output_prev.dropna(subset=['link_call_date']) # drop people who have no previous call under these rules
    output_prev = output_prev.rename(columns={'link_call_date': 'prev_link_call'})
    # merge with full list
    matrix = matrix.merge(output_prev, how='left', on=['client_hash'])
    return matrix

def get_link_screening(matrix, engine, spreadsheet_date):
    client_hash_list = "','".join(matrix['client_hash'])
    client_hash_list = f"'{client_hash_list}'"
    q = f"""
        select
            hashed_mci_uniq_id as client_hash, 
            referral_date as link_call_date
        from clean.link_coordinated_entry
        where hashed_mci_uniq_id in ({client_hash_list})
        and ce_assessment_type='Referral to Prevention Program'
    """
    link_df = pd.read_sql(q, engine, parse_dates=['link_call_date'])
    # find previous call
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['client_hash']]
    output_prev = matrix_tmp.merge(link_df, how='left', on='client_hash')
    output_prev = output_prev.drop(output_prev[output_prev['link_call_date']>spreadsheet_date].index) # drop calls after spreadsheet_date (current date)
    output_prev = output_prev.drop(output_prev[output_prev['link_call_date']<spreadsheet_date-pd.DateOffset(months=6)].index) # drop calls before spreadsheet_date (current date) - 6 months
    output_prev = output_prev.sort_values(by=['link_call_date'], ascending=False)
    output_prev = output_prev.drop_duplicates(subset=['client_hash'])
    output_prev = output_prev.dropna(subset=['link_call_date']) # drop people who have no previous call under these rules
    output_prev = output_prev.rename(columns={'link_call_date': 'prev_link_screen'})
    # merge with full list
    matrix = matrix.merge(output_prev, how='left', on=['client_hash'])
    return matrix

# get PSH/RRH housing program dates for individuals who are currently enrolled in a program 
def get_housing_program_data(matrix, engine, spreadsheet_date):
    mci_ids = matrix.loc[~matrix['mci_uniq_id'].isnull()]['mci_uniq_id']
    mci_id_list = ",".join(mci_ids.astype(int).astype(str))
    q = f"""
        SELECT    
            mci_uniq_id,
            enrollment_start_dt as psh_rrh_start_dt,
            enrollment_end_dt as psh_rrh_end_dt,
            1 as outreach_psh_rrh
        FROM clean.psh_rrh_enrollment pse
        where mci_uniq_id in ({mci_id_list}) and enrollment_end_dt is null
    """
    psh_rrh_df = pd.read_sql(q, engine, parse_dates=['psh_rrh_start_dt', 'psh_rrh_end_dt'])
    # find valid enrollment
    # individual must be CURRENTLY enrolled in PSH/RRH, otherwise we want ACTION to look at the case: start_dt < filingdt and end_dt is null
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['mci_uniq_id']]
    output = matrix_tmp.merge(psh_rrh_df, how='left', on='mci_uniq_id')
    output = output.loc[(output['psh_rrh_start_dt']<spreadsheet_date) & (output['psh_rrh_end_dt'].isnull())]
    output = output.sort_values(by=['psh_rrh_end_dt'], ascending=False)
    output = output.drop_duplicates(subset=['mci_uniq_id'])
    output = matrix.merge(output, how='left', on=['mci_uniq_id'])
    return output

def get_current_cyf(matrix, engine, spreadsheet_date):
    mci_ids = matrix.loc[~matrix['mci_uniq_id'].isnull()]['mci_uniq_id']
    mci_id_list = ",".join(mci_ids.astype(int).astype(str))
    q = f"""
        select
            mci_uniq_id, 
            start_date as cyf_start_dt,
            end_date as cyf_end_dt,
            entity_type as cyf_entity_type,
            entity_id as cyf_entity_id,
            1 as outreach_cyf
        from clean.cyf_ref_case_involvement crci where mci_uniq_id in ({mci_id_list})
    """
    df = pd.read_sql(q, engine, parse_dates=['cyf_start_dt', 'cyf_end_dt'])
    # find valid enrollment
    # individual must be CURRENTLY enrolled with CYF: start_dt < spreadsheet_date and end_dt is null (even if end_dt > spreadsheet_date or filingdt, if the person isn't currently enrolled, they won't have an active caseworker)
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['mci_uniq_id']]
    output = matrix_tmp.merge(df, how='left', on='mci_uniq_id')
    output = output.loc[(output['cyf_start_dt']<spreadsheet_date) & (output['cyf_end_dt'].isnull())]
    # combine
    output = output.sort_values(by=['cyf_end_dt'], ascending=False)
    output = output.drop_duplicates(subset=['mci_uniq_id'])
    output = matrix.merge(output, how='left', on=['mci_uniq_id'])
    return output

def get_client_info(matrix, engine):
    client_hash_list = "','".join(matrix['client_hash'])
    client_hash_list = f"'{client_hash_list}'"
    q = f"""
        SELECT    
            client_hash,
            fname,
            mname,
            lname,
            name_suffix,
            cell_phone,
            home_phone,
            work_phone
        FROM clean.client_feed df
        where client_hash in ({client_hash_list})
    """
    info_df = pd.read_sql(q, engine)
    output = matrix.merge(info_df, how='left', on='client_hash')
    return output

def get_hello_baby_involvement(matrix, engine, spreadsheet_date):
    mci_ids = matrix.loc[~matrix['mci_uniq_id'].isnull()]['mci_uniq_id']
    mci_id_list = ",".join(mci_ids.astype(int).astype(str))
    q = f"""
        SELECT
            mci_uniq_id, 
            contact_date as hello_baby_contact_date,
            priority_provider,
            1 as outreach_ocs
        FROM clean.hello_baby_priority_clients hbc
        where mci_uniq_id in ({mci_id_list})
    """
    output = pd.read_sql(q, engine, parse_dates=['hello_baby_contact_date'])
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['mci_uniq_id']]
    output = matrix_tmp.merge(output, how='left', on='mci_uniq_id')
    output = output.sort_values(by=['hello_baby_contact_date'], ascending=False)
    output = output.drop(output[output['hello_baby_contact_date']<spreadsheet_date-pd.DateOffset(months=6)].index) # drop events outside of the last 6 months
    output = output.drop_duplicates(subset=['mci_uniq_id'])
    output = matrix.merge(output, how='left', on=['mci_uniq_id'])
    return output

def get_family_center_involvement(matrix, engine, spreadsheet_date):
    mci_ids = matrix.loc[~matrix['mci_uniq_id'].isnull()]['mci_uniq_id']
    mci_id_list = ",".join(mci_ids.astype(int).astype(str))
    q = f"""
        SELECT
            mci_uniq_id, 
            contact_date as family_center_contact_date,
            facility_name,
            assigned_worker_name,
            contact_worker_name,
            1 as outreach_ocs
        FROM clean.family_center_clients fcc
        where mci_uniq_id in ({mci_id_list})
    """
    output = pd.read_sql(q, engine, parse_dates=['family_center_contact_date'])
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['mci_uniq_id']]
    output = matrix_tmp.merge(output, how='left', on='mci_uniq_id')
    output = output.sort_values(by=['family_center_contact_date'], ascending=False)
    output = output.drop(output[output['family_center_contact_date']<spreadsheet_date-pd.DateOffset(months=6)].index) # drop events outside of the last 6 months
    output = output.drop_duplicates(subset=['mci_uniq_id'])
    output = matrix.merge(output, how='left', on=['mci_uniq_id'])
    return output

def add_notes_status_rank_history(matrix, engine, spreadsheet_date):
    logging.info(f"SPREADSHEET DATE {spreadsheet_date.strftime('%Y-%m-%d')}")
    # update spreadsheet date 
    matrix['spreadsheet_date'] = spreadsheet_date
    # add entity id
    # acdhs_production.outreach_lists: past cases only -- any new cases will have entity_id already from predictions table
    q = """
        select distinct client_hash, filingdt, entity_id
            from acdhs_production.outreach_lists 
    """
    df_entity_ids = pd.read_sql(q, engine, parse_dates=['filingdt'])
    matrix = matrix.merge(df_entity_ids, how='left', on=['client_hash', 'filingdt'])
    if 'entity_id_x' in matrix.columns:
        # unify entity_id columns
        matrix['entity_id'] = np.where(matrix['entity_id_x'].isnull(), matrix['entity_id_y'],
                                       matrix['entity_id_x'])
    # aggregate written notes -- these are combined regardless of data_source
    # acdhs_production.outreach_lists: past cases only -- notes only apply to cases that were on previous week's list
    q = """
        with distinct_notes as (
        select 
                client_hash, filingdt, as_of_date, spreadsheet_date, data_source, ocn.db_recorded_timestamp, ocn."Attempt 1 notes", ocn."Attempt 2 notes", ocn."Attempt 3 notes",
                rank() over (partition by client_hash, filingdt, as_of_date, spreadsheet_date, data_source order by ocn.db_recorded_timestamp desc) as rn
                from acdhs_production.outreach_contact_notes ocn join acdhs_production.outreach_lists using(client_hash, model_id, as_of_date, spreadsheet_date, prediction_date)
        )
        select * from distinct_notes where rn=1
        order by spreadsheet_date
    """
    df_notes = pd.read_sql(q, engine, parse_dates=['filingdt', 'as_of_date', 'spreadsheet_date'])
    df_notes['prev_notes'] = df_notes[['Attempt 1 notes', 'Attempt 2 notes', 'Attempt 3 notes']].apply(lambda x: '\n- '.join(x.dropna()), axis=1)
    df_notes['prev_notes'] = df_notes['prev_notes'].str.strip()
    df_notes = df_notes.dropna(subset=['prev_notes'])
    df_notes = df_notes.loc[df_notes['prev_notes']!='']
    df_notes['prev_notes'] = '- ' + df_notes['prev_notes'].astype(str)
    out = df_notes.groupby(['client_hash', 'filingdt'], as_index=False).agg(previous_notes=('prev_notes', '\n'.join))
    matrix = matrix.merge(out, how='left', on=['client_hash', 'filingdt'])
    # get most recent status per outreach org
    # acdhs_production.outreach_lists: past cases only -- status only available from existing sheets, not the new sheet we're generating
    q = """
        with status as (
        select 
                client_hash, filingdt, as_of_date, spreadsheet_date, data_source, ocn.db_recorded_timestamp, "Contact status", "Court status", "Rental assistance status", "Did this help tenant?",
                rank() over (partition by client_hash, filingdt, data_source order by spreadsheet_date desc, ocn.db_recorded_timestamp desc) as rn
                from acdhs_production.outreach_contact_notes ocn join acdhs_production.outreach_lists using(client_hash, model_id, as_of_date, spreadsheet_date, prediction_date)
        )
        select * from status
        where rn=1
    """
    df_cats = pd.read_sql(q, engine, parse_dates=['filingdt', 'as_of_date', 'spreadsheet_date'])
    df_cats_ids = df_cats.copy()
    df_cats_ids = df_cats_ids[['client_hash', 'filingdt']]
    df_cats_ids.drop_duplicates(inplace=True)
    for org in orgs:
        # make status columns for each org
        df_cats_org = df_cats.loc[df_cats['data_source']==org][["client_hash", "filingdt", "as_of_date", "spreadsheet_date", "Contact status", "Court status", "Rental assistance status", "Did this help tenant?"]]
        for status_col in status_columns:
            df_cats_org[f"{org} {status_col}"] = df_cats_org[status_col]
        df_cats_org = df_cats_org.drop(['as_of_date', "spreadsheet_date", "Contact status", 'Court status', "Rental assistance status", "Did this help tenant?"], axis=1)
        df_cats_ids = df_cats_ids.merge(df_cats_org, how='left', on=['client_hash', 'filingdt'])
    matrix = matrix.merge(df_cats_ids, how='left', on=['client_hash', 'filingdt'])
    # get most recent rank
    # acdhs_production.outreach_lists: past cases only -- most recent rank either comes from previous sheet (week) or from predictions (rank_abs_no_ties and score_x/y below)
    q = """
        with rank_table as (
            select 
                client_hash, 
                filingdt, 
                as_of_date, 
                spreadsheet_date, 
                coalesce(most_recent_rank, rank) as most_recent_rank,
                score,
                rank() over (partition by client_hash, filingdt order by spreadsheet_date desc) as rn
            from acdhs_production.outreach_lists 
        )
        select client_hash, filingdt, score, most_recent_rank from rank_table where rn=1;
    """
    df_rank = pd.read_sql(q, engine, parse_dates=['filingdt', 'as_of_date', 'spreadsheet_date'])
    matrix = matrix.merge(df_rank, how='left', on=['client_hash', 'filingdt'])
    if 'score_x' in matrix.columns: # we took some cases from predict_forward
        # if case wasn't on a previous list, take rank from predictions table
        matrix['most_recent_rank'] = np.where(matrix['most_recent_rank'].isnull(), matrix['rank_abs_no_ties'],
                                              matrix['most_recent_rank'])
        # unify score columns
        matrix['score'] = np.where(matrix['score_x'].isnull(), matrix['score_y'],
                                       matrix['score_x'])
    # get history info
    q = f"""
        select 
            client_hash,
            filingdt,
            spreadsheet_date,
            assign_action,
            coalesce(lag(assign_action, 1) over (partition by client_hash, filingdt order by spreadsheet_date), 0) as prev_assign_action,
            assign_cyf,
            coalesce(lag(assign_cyf, 1) over (partition by client_hash, filingdt order by spreadsheet_date), 0) as prev_assign_cyf,
            assign_ocs,
            coalesce(lag(assign_ocs, 1) over (partition by client_hash, filingdt order by spreadsheet_date), 0) as prev_assign_ocs,
            assign_mdj,
            coalesce(lag(assign_mdj, 1) over (partition by client_hash, filingdt order by spreadsheet_date), 0) as prev_assign_mdj,
            assign_renthelp,
            coalesce(lag(assign_renthelp, 1) over (partition by client_hash, filingdt order by spreadsheet_date), 0) as prev_assign_renthelp,
            assign_psh_rrh,
            coalesce(lag(assign_psh_rrh, 1) over (partition by client_hash, filingdt order by spreadsheet_date), 0) as prev_assign_psh_rrh
        from acdhs_production.outreach_lists
    """
    df_history_tracker = pd.read_sql(q, engine, parse_dates=['filingdt', 'spreadsheet_date'])
    df_history_copy = df_history_tracker.copy()
    df_history_copy = df_history_copy.loc[df_history_copy['client_hash'].isin(matrix['client_hash'])]
    df_history_copy.sort_values(by=['spreadsheet_date'])
    df_history_last = df_history_copy.groupby(['client_hash']).last() # the most recent list assignments (before today)
    df_history_last.drop(['prev_assign_action', 'prev_assign_cyf', 'prev_assign_ocs', 'prev_assign_mdj', 'prev_assign_renthelp', 'prev_assign_psh_rrh'], axis=1, inplace=True)
    df_history_last.rename(columns={'assign_action': 'prev_assign_action', 'assign_cyf': 'prev_assign_cyf', 'assign_ocs': 'prev_assign_ocs', 'assign_mdj': 'prev_assign_mdj', 'assign_renthelp': 'prev_assign_renthelp', 'assign_psh_rrh': 'prev_assign_psh_rrh'}, inplace=True)
    df_history_last.drop(['filingdt', 'spreadsheet_date'], axis=1, inplace=True)
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['client_hash', 'filingdt', 'spreadsheet_date', 'assign_ACTION', 'assign_CYF', 'assign_OCS', 'assign_MDJ', 'assign_RentHelp', 'assign_PSH_RRH']]
    matrix_tmp.rename(columns={'assign_ACTION': 'assign_action', 'assign_CYF': 'assign_cyf', 'assign_OCS': 'assign_ocs', 'assign_MDJ': 'assign_mdj', 'assign_RentHelp': 'assign_renthelp', 'assign_PSH_RRH': 'assign_psh_rrh'}, inplace=True)
    df_history_fill = matrix_tmp.merge(df_history_last, how='left', on=['client_hash'])
    df_history_fill.fillna(0, inplace=True)
    df_history_tracker = pd.concat([df_history_tracker, df_history_fill])
    df_history = pd.DataFrame()
    for org in orgs:
        # get addition/removal strings for this org
        df_org_changes = df_history_tracker.loc[df_history_tracker[f"assign_{org.lower()}"]!=df_history_tracker[f"prev_assign_{org.lower()}"]]
        df_org_changes = df_org_changes.rename(columns={'spreadsheet_date': 'event_date'})
        df_org_changes['str_desc'] = df_org_changes.apply(get_addition_removal_str, org=org, axis=1)
        df_org_changes.drop(['assign_action', 'prev_assign_action', 'assign_cyf', 'prev_assign_cyf', 'assign_ocs', 'prev_assign_ocs', 'assign_mdj', 'prev_assign_mdj', 'assign_renthelp', 'prev_assign_renthelp', 'assign_psh_rrh', 'prev_assign_psh_rrh'], axis=1, inplace=True)
        df_history = pd.concat([df_history, df_org_changes])
    # add first incidence of (client_hash, filingdt) on list
    q = """
        with a as(
        SELECT ROW_NUMBER() OVER (PARTITION BY client_hash, filingdt ORDER BY spreadsheet_date) as rn, *
            FROM acdhs_production.outreach_lists 
            )
        select client_hash, filingdt, spreadsheet_date as event_date, concat('Appeared on list on ', spreadsheet_date) as str_desc
        from a 
        where rn = 1
    """
    df_history = pd.concat([df_history, pd.read_sql(q, engine, parse_dates=['filingdt', 'event_date'])])
    # add current incident for (client_hash, filingdt) that don't appear in existing acdhs_production.outreach_lists
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['client_hash', 'filingdt']]
    df_history_tmp = df_history.copy()[['client_hash', 'filingdt']]
    df_incidence_add = matrix_tmp.merge(df_history_tmp.drop_duplicates(), on=['client_hash','filingdt'], how='left', indicator=True)
    df_incidence_add = df_incidence_add.loc[df_incidence_add['_merge']=='left_only'] # get cases that don't exist in history
    df_incidence_add['event_date'] = spreadsheet_date
    df_incidence_add['str_desc'] = f'Appeared on list on ' + spreadsheet_date.strftime('%Y-%m-%d') 
    df_incidence_add.drop(['_merge'], axis=1, inplace=True)
    df_history = pd.concat([df_history, df_incidence_add])
    # add evictions
    q = """
        select distinct client_hash, filingdt from acdhs_production.outreach_lists
    """
    df_filing = pd.read_sql(q, engine, parse_dates=['filingdt'])
    matrix_tmp = matrix.copy()
    matrix_tmp = matrix_tmp[['client_hash', 'filingdt']]
    df_filing_tmp = df_filing.copy()[['client_hash', 'filingdt']]
    df_filing_add = matrix_tmp.merge(df_filing_tmp.drop_duplicates(), on=['client_hash','filingdt'], how='left', indicator=True)
    df_filing_add = df_filing_add.loc[df_filing_add['_merge']=='left_only'] # get cases that don't exist in history
    df_filing = pd.concat([df_filing, df_filing_add])
    df_filing['event_date'] = df_filing['filingdt']
    df_filing['str_desc'] = f'Eviction filed ' + df_filing['filingdt'].dt.strftime('%Y-%m-%d')
    df_filing.drop(['_merge'], axis=1, inplace=True)
    df_history = pd.concat([df_history, df_filing])
    # add status changes
    # acdhs_production.outreach_lists: past cases only -- status only available from existing sheets, not the new sheet we're generating
    q = """
        with distinct_statuses as (
        select 
                client_hash, filingdt, spreadsheet_date, data_source, ocn.db_recorded_timestamp, ocn."Contact status", ocn."Court status", ocn."Rental assistance status", ocn."Did this help tenant?",
                rank() over (partition by client_hash, filingdt, spreadsheet_date, data_source order by ocn.db_recorded_timestamp asc) as rn
                from acdhs_production.outreach_contact_notes ocn join acdhs_production.outreach_lists using(client_hash, model_id, spreadsheet_date, prediction_date)
        ),
        statuses as (
                select 
                    client_hash,
                    filingdt,
                    spreadsheet_date,
                    data_source,
                    db_recorded_timestamp,
                    coalesce("Contact status", '') as "Contact status",
                    coalesce("Court status", '') as "Court status",
                    coalesce("Rental assistance status", '') as "Rental assistance status"
                from distinct_statuses where rn=1
                order by spreadsheet_date, db_recorded_timestamp
        ),
        status_changes as
        (
            select 
                distinct on (client_hash, db_recorded_date, data_source)
                client_hash,
                filingdt,
                spreadsheet_date, 
                date(db_recorded_timestamp) as db_recorded_date,
                data_source,
                "Contact status",
                "Court status",
                "Rental assistance status",
                coalesce(lag("Contact status", 1) over (partition by client_hash, data_source order by spreadsheet_date),'') as prev_contact_status,
                coalesce(lag("Court status", 1) over (partition by client_hash, data_source order by spreadsheet_date),'') as prev_court_status,
                coalesce(lag("Rental assistance status", 1) over (partition by client_hash, data_source order by spreadsheet_date),'') as prev_ra_status
            from statuses
        )  
        select 
            client_hash,
            filingdt,
            db_recorded_date - '2days'::interval as event_date,
            case when "Contact status"!=prev_contact_status then concat('Contact status changed to "',"Contact status",'" by ',data_source, ' on ',(db_recorded_date - '2days'::interval)::date) else '' end as str_desc
        from status_changes
        where "Contact status" != prev_contact_status and "Contact status" != '' and "Contact status" != '--'
        union
            select 
                client_hash,
                filingdt,
                db_recorded_date - '2days'::interval as event_date,
                case when "Court status"!=prev_court_status then concat('Court status changed to "',"Court status",'" by ',data_source, ' on ',(db_recorded_date - '2days'::interval)::date) else '' end as str_desc
            from status_changes
            where "Court status" != prev_court_status and "Court status" != '' and "Court status" != '--'
        union
            select 
                client_hash,
                filingdt,
                db_recorded_date - '2days'::interval as event_date,
                case when "Rental assistance status"!=prev_ra_status then concat('Rental assistance status changed to "',"Rental assistance status",'" by ',data_source, ' on ',(db_recorded_date - '2days'::interval)::date) else '' end as str_desc
            from status_changes
            where "Rental assistance status" != prev_ra_status and "Rental assistance status" != '' and "Rental assistance status" != '--'
    """
    df_status_change = pd.read_sql(q, engine, parse_dates=['filingdt', 'event_date'])
    df_history = pd.concat([df_history, df_status_change])
    df_history['str_len'] = df_history['str_desc'].str.len()
    df_history.sort_values(by=['event_date', 'str_len'], ascending=[True, True], inplace=True)
    df_history.drop(['str_len'], axis=1, inplace=True)
    # concat message strings ordered by event_date
    out = df_history.groupby(['client_hash'], as_index=False).agg(history_desc=('str_desc', '\n'.join))
    matrix = matrix.merge(out, how='left', on=['client_hash'])
    matrix = matrix.sort_values(by=['most_recent_rank'])
    # add first appearance info
    # acdhs_production.outreach_lists: past cases only -- assumes sum=0 if this is the first appearance
    client_hash_list = "','".join(matrix['client_hash'])
    client_hash_list = f"'{client_hash_list}'"
    q = f"""
    select 
        client_hash,
        case when sum(assign_action)=0 then 1 else 0 end as first_action_appearance,
        case when sum(assign_cyf)=0 then 1 else 0 end as first_cyf_appearance,
        case when sum(assign_mdj)=0 then 1 else 0 end as first_mdj_appearance,
        case when sum(assign_ocs)=0 then 1 else 0 end as first_ocs_appearance,
        case when sum(assign_renthelp)=0 then 1 else 0 end as first_renthelp_appearance,
        case when sum(assign_psh_rrh)=0 then 1 else 0 end as first_psh_rrh_appearance
    from acdhs_production.outreach_lists where client_hash in ({client_hash_list}) and spreadsheet_date<'{spreadsheet_date}'::date
    group by client_hash
    """
    df = pd.read_sql(q, engine)
    # add brand new cases
    new_cases = matrix.copy()
    new_cases = new_cases.loc[~matrix['client_hash'].isin(df['client_hash'])]
    for org in orgs:
        new_cases[f"first_{org.lower()}_appearance"] = new_cases[f"assign_{org}"]
    new_cases = new_cases[['client_hash', 'first_action_appearance', 'first_cyf_appearance', 'first_mdj_appearance', 'first_ocs_appearance', 'first_renthelp_appearance', 'first_psh_rrh_appearance']]
    df = pd.concat([df, new_cases])
    for row in df.itertuples():
        if row.first_action_appearance==1:
            first_time_names['ACTION'].add(row.client_hash)
        if row.first_cyf_appearance==1:
            first_time_names['CYF'].add(row.client_hash)
        if row.first_mdj_appearance==1:
            first_time_names['MDJ'].add(row.client_hash)
        if row.first_ocs_appearance==1:
            first_time_names['OCS'].add(row.client_hash)
        if row.first_renthelp_appearance==1:
            first_time_names['RentHelp'].add(row.client_hash)
        if row.first_psh_rrh_appearance==1:
            first_time_names['PSH_RRH'].add(row.client_hash)
    return matrix

def get_addition_removal_str(row, org):
    if row[f'assign_{org.lower()}'] == 1:
        return f"Added to {org}'s outreach list on {row.event_date.strftime('%Y-%m-%d')}"
    else:
        return f"Removed from {org}'s outreach list on {row.event_date.strftime('%Y-%m-%d')}"

def format_spreadsheet(matrix, score_round_decimals=2):
    matrix = matrix.round({'score': score_round_decimals})
    # add notes columns
    matrix["Attempt 1 notes"] = " "
    matrix["Attempt 2 notes"] = " "
    matrix["Attempt 3 notes"] = " "
    matrix.loc[matrix['city_of_pgh']==False, 'city_of_pgh'] = None
    # column order
    status_cols = []
    for org in orgs:
        for status_col in status_columns:
            col_name = f"{org} {status_col}"
            status_cols.append(col_name)
    id_cols = ['mci_uniq_id', 'fname','mname','lname','name_suffix'] + status_cols + ['cell_phone','home_phone','work_phone', 'previous_notes', 'as_of_date', 'spreadsheet_date', "Attempt 1 notes", "Attempt 2 notes", "Attempt 3 notes", "history_desc"]
    outreach_cols = ['outreach_list', 'assign_PSH_RRH', 'psh_rrh_start_dt', 'psh_rrh_end_dt',
                     'assign_CYF', 'cyf_start_dt', 'cyf_end_dt', 'cyf_entity_type', 'cyf_entity_id',
                     'assign_RentHelp',
                     'assign_MDJ',
                     'assign_OCS', 'hello_baby_contact_date', 'priority_provider', 'family_center_contact_date', 'facility_name', 'assigned_worker_name', 'contact_worker_name',
                     'assign_ACTION']
    known_cols = ['filingdt', 
                  'matter_id', 
                  'docketno',
                  'dispositiondt',
                  'city_of_pgh',
                   'arrears_only', 
                    'judgment_for_landlord', 
                    'claimamount',
                    'totaljudgmentamount',
            'entry_of_satisfaction_issue_dt',
            'entry_of_satisfaction_request_dt',
                   'ofp_issue_dt', 
                  'prev_link_call',
                  'prev_link_screen',
                  'prev_ra_app_update',
                  'prev_ra_app_submitted',
                  'prev_ra_app_tnumber',
                  'prev_ra_payment',
                  'prev_ra_payment_tnumber',
                  'erap_months']
    meta = ['client_hash', 'model_id', 'entity_id', 'prediction_date']
    future_cols = ['dod']
    cols_to_keep = id_cols + outreach_cols + known_cols + future_cols + meta
    cols_to_drop = []
    for col_name in matrix.columns:
        if col_name not in cols_to_keep:
            cols_to_drop.append(col_name)
    logging.info(f"columns dropped from output spreadsheet: {', '.join(cols_to_drop)}")
    matrix = matrix.drop(cols_to_drop, axis=1)
    matrix = matrix[cols_to_keep]
    date_cols = ['as_of_date', 'spreadsheet_date', 'prediction_date', 'filingdt', 'dispositiondt', 'ofp_issue_dt', 'prev_ra_payment', 'psh_rrh_start_dt', 'psh_rrh_end_dt', 'cyf_start_dt', 'cyf_end_dt', 'hello_baby_contact_date', 'family_center_contact_date', 'prev_link_call', 'prev_link_screen', 'prev_ra_app_update', 'prev_ra_app_submitted']
    for col in date_cols:
        matrix[col] = pd.to_datetime(matrix[col]).dt.date
    # set empty values to "Not Known Yet"
    matrix.loc[matrix['dispositiondt'].isna(), 'dispositiondt'] = 'Not Known Yet'
    matrix.loc[matrix['arrears_only']=='', 'arrears_only'] = 'Not Known Yet'
    matrix.loc[matrix['judgment_for_landlord']=='', 'judgment_for_landlord'] = 'Not Known Yet'
    matrix.loc[matrix['totaljudgmentamount'].isna(), 'totaljudgmentamount'] = 'Not Known Yet'
    # format phone numbers
    matrix['cell_phone'].fillna(0, inplace=True)
    matrix['cell_phone'].replace('', 0, inplace=True)
    matrix['cell_phone'] = matrix['cell_phone'].astype(int).astype(str)
    matrix['cell_phone'].replace('0', '', inplace=True)
    matrix['work_phone'] = matrix['work_phone'].fillna(0)
    matrix['work_phone'].replace('', 0, inplace=True)
    matrix['work_phone'] = matrix['work_phone'].astype(int).astype(str)
    matrix['work_phone'].replace('0', '', inplace=True)
    matrix['home_phone'].fillna(0, inplace=True)
    matrix['home_phone'].replace('', 0, inplace=True)
    matrix['home_phone'] = matrix['home_phone'].astype(int).astype(str)
    matrix['home_phone'].replace('0', '', inplace=True)
    return matrix

def get_death_dates(matrix, engine):
    client_hash_list = "','".join(matrix['client_hash'])
    client_hash_list = f"'{client_hash_list}'"
    q = f"""
        SELECT    
            client_hash,
            dod
        FROM clean.client_feed df
        where client_hash in ({client_hash_list})
    """
    dod_df = pd.read_sql(q, engine, parse_dates=['dod'])
    output = matrix.merge(dod_df, how='left', on='client_hash')
    return output

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Generate formatted outreach spreadsheets from model predictions")

    parser.add_argument(
        "-m",
        "--model_id",
        type=int,
        help='triage model id that has generated predictions in acdhs_production (int)',
        required=True
    )
    
    parser.add_argument(
        "-pfd",
        "--pfwd_as_of_date",
        type=str,
        help='Predict forward as_of_date (used to retrieve new cases from most recent list when needed)',
        required=True
    )

    parser.add_argument("--write_to_db", action="store_true", 
                    help="write list to db") 
    
    parser.add_argument("--rct", action="store_true", 
                    help="fill from predict_forward list according to RCT records") 
    
    parser.add_argument(
        "-msd",
        "--monday_spreadsheet_date",
        type=str,
        required=True
    )
    parser.add_argument(
        "--random_seed",
        type=int,
        required=True
    )
    parser.add_argument(
        "--sample_seed1",
        type=int,
        required=True
    )
    parser.add_argument(
        "--sample_seed2",
        type=int,
        required=True
    )
    parser.add_argument(
        "--sample_seed3",
        type=int,
        required=True
    )

    args = parser.parse_args()
    if args.write_to_db:
        args.write_to_db = True
    else:
        args.write_to_db = False
    if args.rct:
        args.rct = True
        rct_str = "_rct"
    else:
        args.rct = False
        rct_str = ""
    sample_seeds = [args.sample_seed1, args.sample_seed2, args.sample_seed3]
    args.spreadsheet_date = datetime.datetime.today().strftime('%Y-%m-%d')
    out_dir = f"/mnt/data/projects/acdhs_housing/outreach_spreadsheets/{args.monday_spreadsheet_date}"
    filename = f"list_{args.monday_spreadsheet_date}_pfd{args.pfwd_as_of_date.replace('-','')}_m{args.model_id}{rct_str}_{args.random_seed}_{args.sample_seed1}_{args.sample_seed2}_{args.sample_seed3}.xlsx"
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    filepath = os.path.join(out_dir, filename)
    logging.info(f"file location: {filepath}")

    if os.path.exists(filepath):
        logging.error("ERROR: file exists, stopping")
        sys.exit(0)
    df = generate_spreadsheet(args.spreadsheet_date, args.pfwd_as_of_date, args.model_id, filepath, write_to_db=args.write_to_db, rct=args.rct, random_seed=args.random_seed, sample_seeds=sample_seeds)
